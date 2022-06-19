import openpyxl

class HomePageData:

        test_HomePage_data = [{"firstname":"rahul", "lastname":"shetty", "gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

        @staticmethod
        def getTestData(test_case_name):
                book = openpyxl.load_workbook("C:\\Users\\xcnh76\Desktop\PythonDemo.xlsx")
                sheet = book.active
                Dict = {}
                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == "Testcase2":

                                for j in range(2,
                                               sheet.max_column + 1):  # runs inner loop until it is finished, and comes back to first one

                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return[Dict]  # sending dictionary in the list, because it has to
